"""Preview-first CSV/XLSX Catalogue maintenance importer."""

from __future__ import annotations

import csv
import re
from pathlib import Path

from django.db import transaction

from apps.core.enums import ActivityType, SsaIntervention

from .models import (
    ActivityCatalogueItem,
    ActivityEligibilityRule,
    ActivityInterventionMapping,
    CatalogueActivityType,
    CatalogueStatus,
    DeliveryMethod,
    MappingMode,
)
from .seeding import normalize_alias
from .services import create_version


REQUIRED_COLUMNS = {
    "Stable Code",
    "Activity Name",
    "Activity Type",
    "Delivery Method",
    "Intervention Mapping",
    "Mapping Mode",
    "Target Audience",
    "Staff Delivery",
    "Partner Delivery",
    "Costing Profile",
    "Evidence Profile",
    "Salesforce Record Type",
    "Status",
}

TYPE_VALUES = {
    normalize_alias(label): value for value, label in CatalogueActivityType.choices
}
TYPE_VALUES.update({value: value for value in CatalogueActivityType.values})
DELIVERY_VALUES = {
    normalize_alias(label): value for value, label in DeliveryMethod.choices
}
DELIVERY_VALUES.update({value: value for value in DeliveryMethod.values})
DELIVERY_VALUES.update(
    {
        "visit": DeliveryMethod.SCHOOL_VISIT,
        "school visits": DeliveryMethod.SCHOOL_VISIT,
    }
)
MODE_VALUES = {normalize_alias(label): value for value, label in MappingMode.choices}
MODE_VALUES.update({value: value for value in MappingMode.values})
STATUS_VALUES = {
    normalize_alias(label): value for value, label in CatalogueStatus.choices
}
STATUS_VALUES.update({value: value for value in CatalogueStatus.values})
INTERVENTION_VALUES = {
    normalize_alias(label): value for value, label in SsaIntervention.choices
}
INTERVENTION_VALUES.update(
    {normalize_alias(value): value for value in SsaIntervention.values}
)
INTERVENTION_VALUES.update(
    {
        "christ like behaviour": SsaIntervention.CHRISTLIKE_BEHAVIOUR,
        "enrollment": SsaIntervention.ENROLMENT,
        "teachers environment": SsaIntervention.TEACHING_ENVIRONMENT,
        "teacher s environment": SsaIntervention.TEACHING_ENVIRONMENT,
        "government requirements": SsaIntervention.GOVERNMENT_REQUIREMENT,
    }
)

WORKFLOW_BY_SHAPE = {
    (
        CatalogueActivityType.SCHOOL_VISIT,
        DeliveryMethod.SCHOOL_VISIT,
    ): ActivityType.SCHOOL_VISIT,
    (
        CatalogueActivityType.TRAINING,
        DeliveryMethod.IN_SCHOOL_TRAINING,
    ): ActivityType.IN_SCHOOL_TRAINING,
    (
        CatalogueActivityType.TRAINING,
        DeliveryMethod.CLUSTER_TRAINING,
    ): ActivityType.CLUSTER_TRAINING,
    (
        CatalogueActivityType.TRAINING,
        DeliveryMethod.CLUSTER_MEETING,
    ): ActivityType.CLUSTER_MEETING,
    (CatalogueActivityType.TRAINING, DeliveryMethod.ONLINE): ActivityType.TRAINING,
    (
        CatalogueActivityType.TRAINING,
        DeliveryMethod.SCHOOL_VISIT,
    ): ActivityType.BASELINE_SSA_VISIT,
    (CatalogueActivityType.YOUTH_CAMP, DeliveryMethod.GROUP): ActivityType.TRAINING,
    (CatalogueActivityType.ADMIN, DeliveryMethod.ADMIN): ActivityType.PARTNER_ACTIVITY,
}


def _truth(value):
    return str(value or "").strip().casefold() in {"1", "true", "yes", "y", "allowed"}


def _rows(path: Path):
    if path.suffix.casefold() == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as stream:
            yield from csv.DictReader(stream)
        return
    if path.suffix.casefold() not in {".xlsx", ".xlsm"}:
        raise ValueError("Catalogue import supports CSV or XLSX files.")
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    values = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(values)]
    for values_row in values:
        yield dict(zip(headers, values_row))


def _normalize_row(raw, row_number):
    errors = []
    code = str(raw.get("Stable Code") or "").strip().upper()
    if not re.fullmatch(r"[A-Z][A-Z0-9_]{2,95}", code):
        errors.append(
            "Stable Code must use uppercase letters, numbers and underscores."
        )
    name = str(raw.get("Activity Name") or "").strip()
    if not name:
        errors.append("Activity Name is required.")
    activity_type = TYPE_VALUES.get(
        normalize_alias(str(raw.get("Activity Type") or ""))
    )
    delivery = DELIVERY_VALUES.get(
        normalize_alias(str(raw.get("Delivery Method") or ""))
    )
    mode = MODE_VALUES.get(normalize_alias(str(raw.get("Mapping Mode") or "")))
    status = STATUS_VALUES.get(normalize_alias(str(raw.get("Status") or "")))
    intervention_raw = str(raw.get("Intervention Mapping") or "").strip()
    intervention = (
        INTERVENTION_VALUES.get(normalize_alias(intervention_raw))
        if intervention_raw
        else None
    )
    for value, label in (
        (activity_type, "Activity Type"),
        (delivery, "Delivery Method"),
        (mode, "Mapping Mode"),
        (status, "Status"),
    ):
        if not value:
            errors.append(f"{label} is not recognized.")
    fixed = mode in {MappingMode.FIXED, MappingMode.MULTIPLE_ALLOWED}
    if fixed and not intervention:
        errors.append("A fixed mapping requires one canonical SSA intervention.")
    if not fixed and intervention:
        errors.append(
            "Dynamic, prerequisite and administrative mappings cannot contain an SSA intervention."
        )
    workflow = WORKFLOW_BY_SHAPE.get((activity_type, delivery))
    if activity_type and delivery and not workflow:
        errors.append(
            "Activity Type and Delivery Method have no approved workflow shape."
        )
    cluster = delivery in {
        DeliveryMethod.CLUSTER_MEETING,
        DeliveryMethod.CLUSTER_TRAINING,
    }
    return {
        "row": row_number,
        "stable_code": code,
        "source_name": name,
        "display_name": name,
        "activity_type": activity_type,
        "delivery_method": delivery,
        "workflow_kind": workflow,
        "mapping_mode": mode,
        "intervention": intervention,
        "target_audience": str(raw.get("Target Audience") or "").strip(),
        "staff_delivery_allowed": _truth(raw.get("Staff Delivery")),
        "partner_delivery_allowed": _truth(raw.get("Partner Delivery")),
        "individual_school_allowed": not cluster
        and activity_type != CatalogueActivityType.ADMIN,
        "cluster_delivery_allowed": cluster,
        "project_delivery_allowed": activity_type != CatalogueActivityType.ADMIN,
        "requires_school": not cluster and activity_type != CatalogueActivityType.ADMIN,
        "requires_cluster": cluster,
        "requires_project": False,
        "requires_current_ssa": mode
        not in {
            MappingMode.ADMINISTRATIVE,
            MappingMode.SSA_COMPLETION_PREREQUISITE,
        },
        "requires_source_activity": mode == MappingMode.INHERIT_FROM_SOURCE_ACTIVITY,
        "costing_profile": str(raw.get("Costing Profile") or "").strip(),
        "evidence_profile": str(raw.get("Evidence Profile") or "").strip(),
        "salesforce_record_type": str(raw.get("Salesforce Record Type") or "")
        .strip()
        .upper(),
        "status": status,
        "errors": errors,
    }


@transaction.atomic
def import_catalogue(path, *, actor_id, commit=False):
    path = Path(path)
    raw_rows = list(_rows(path))
    headers = set(raw_rows[0].keys()) if raw_rows else set()
    missing = sorted(REQUIRED_COLUMNS - headers)
    if missing:
        return {
            "valid": False,
            "missingColumns": missing,
            "rows": [],
            "committed": False,
        }
    rows = [_normalize_row(row, index) for index, row in enumerate(raw_rows, 2)]
    seen = set()
    for row in rows:
        if row["stable_code"] in seen:
            row["errors"].append("Duplicate Stable Code in import file.")
        seen.add(row["stable_code"])
    valid = not any(row["errors"] for row in rows)
    if valid and commit:
        for normalized in rows:
            row = dict(normalized)
            mapping_mode = row.pop("mapping_mode")
            intervention = row.pop("intervention")
            row.pop("row")
            row.pop("errors")
            stable_code = row.pop("stable_code")
            existing = ActivityCatalogueItem.objects.filter(
                stable_code=stable_code
            ).first()
            item, _created = ActivityCatalogueItem.objects.update_or_create(
                stable_code=stable_code,
                defaults={
                    **row,
                    "updated_by": actor_id,
                    **({"created_by": actor_id} if existing is None else {}),
                },
            )
            mapping, _ = ActivityInterventionMapping.objects.update_or_create(
                catalogue_item=item,
                intervention=intervention,
                mapping_mode=mapping_mode,
                defaults={"active": True, "is_primary": True, "priority": 100},
            )
            ActivityInterventionMapping.objects.filter(catalogue_item=item).exclude(
                id=mapping.id
            ).update(active=False)
            ActivityEligibilityRule.objects.get_or_create(catalogue_item=item)
            create_version(
                item,
                actor_id=actor_id,
                reason="Authorized Catalogue spreadsheet import.",
            )
        from apps.audit.services import log as audit_log

        audit_log(
            action="activity_catalogue_import",
            subject_kind="ActivityCatalogue",
            subject_id=path.name,
            actor_id=actor_id,
            payload={"rows": len(rows), "file": path.name},
        )
    if not commit:
        transaction.set_rollback(True)
    return {
        "valid": valid,
        "missingColumns": [],
        "rows": rows,
        "committed": bool(valid and commit),
    }
