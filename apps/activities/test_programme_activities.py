"""Contract tests for Non-School Programme Activities (2026-07 feature).

One entry service (planning.schedule_programme_activity) gated by the
planning.manualActivity.create permission, feeding the SAME activities.create
funnel every school activity uses: catalogue governance, REG-02 scheduling
policy, funded-scheduling gate, §9 cross-month cost allocation, advance +
weekly-fund-request channels, My Plan, cancel/reschedule, §35 health checks,
and the double-click duplicate guard under the StaffProfile row lock.
"""

from __future__ import annotations

import threading
from io import BytesIO
from datetime import date

from django.db import connections
from django.test import TestCase, TransactionTestCase
from openpyxl import load_workbook

from apps.accounts.models import StaffProfile, User
from apps.activities import services as asvc
from apps.activities.models import Activity, ActivityScheduleCostLine
from apps.activities.work_plan_health import work_plan_health
from apps.budget.models import CostSetting
from apps.core.exceptions import BadRequest, Forbidden
from apps.fund_requests.models import (
    AdvanceRequest,
    WeeklyFundRequest,
    WeeklyFundRequestLine,
)
from apps.geography.models import District, Region
from apps.my_plan import services as my_plan
from apps.notifications.models import Notification
from apps.planning import services as psvc
from apps.projects.models import Project

# Canonical programme component rates seeded by apps.budget.reference (UGX).
# Conferences and student camps use the dedicated CD programme-event rate
# family, separate from training and cluster-meeting participant costs.
VENUE = 300_000
MEAL = 15_000
FACILITATION = 100_000
TRANSPORT = 100_000
MATERIALS = 5_000
ACCOMMODATION = 80_000

THREE_DAY_TOTAL = (
    3 * VENUE
    + 40 * 3 * MEAL
    + 3 * FACILITATION
    + 3 * TRANSPORT
    + 40 * MATERIALS
    + 2 * ACCOMMODATION
)

FY = "2026"  # Aug/Sep 2026 both sit in operational FY2026 (Oct 1 boundary)

# 2026 calendar facts the fixtures below rely on.
MON_AUG_31 = "2026-08-31T09:00:00+03:00"
TUE_SEP_1 = "2026-09-01T09:00:00+03:00"
FRI_SEP_4 = "2026-09-04T09:00:00+03:00"
MON_SEP_14 = "2026-09-14T09:00:00+03:00"
SUN_AUG_30 = "2026-08-30T09:00:00+03:00"  # a Sunday (REG-02)

_OMIT = object()


def _person(key: str, role: str) -> User:
    user = User.objects.create_user(
        email=f"prog-{key}@edify.org",
        name=f"Prog {key}",
        roles=[role],
        active_role=role,
        password="x",
        is_active=True,
    )
    StaffProfile.objects.create(user=user, title=role)
    return user


def _payload(**over) -> dict:
    """A well-formed programme payload; pass field=_OMIT to drop a key.

    Uses the governed catalogue's own Group-delivered item — the real
    non-school programme activities are the source catalogue's camps and
    conferences, not a parallel invented set."""
    data = {
        "catalogueItemId": "STUDENT_CONFERENCE_CAMPS",
        "scheduledDate": MON_AUG_31,
        "supportRationale": "organizational_priority",
        "expectedParticipants": 40,
        "plannedSchoolCount": 12,
        "programmeActivityType": "student_activities",
        "programmeDeliveryMode": "group",
        "focusIntervention": "christlike_behaviour",
        "venue": "Kampala Conference Centre",
        "activityPurposeText": "Annual leadership conference",
    }
    data.update(over)
    return {k: v for k, v in data.items() if v is not _OMIT}


def _schedule(principal, **over) -> Activity:
    result = psvc.schedule_programme_activity(_payload(**over), principal)
    return Activity.objects.get(id=result["id"])


class _ProgrammeFixture(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.cceo = _person("cceo", "CCEO")
        cls.pl = _person("pl", "Program Lead")
        cls.region = Region.objects.create(name="Prog Region")
        cls.district = District.objects.create(
            name="Prog District", region=cls.region, district_type="primary"
        )
        cls.project = Project.objects.create(name="SP Prog EdTech", category="pilot")


# ── Creation contract ────────────────────────────────────────────────────────
class ProgrammeActivityCreateTest(_ProgrammeFixture):
    def test_authorized_staff_can_create_non_school_activity(self):
        a = _schedule(self.cceo)
        self.assertEqual(a.planning_source, "manual_work_plan")
        self.assertEqual(a.planning_source, "manual_work_plan")
        self.assertEqual(a.activity_context_type, "organization")
        self.assertEqual(a.status, "scheduled")
        self.assertEqual(a.support_rationale, "organizational_priority")
        self.assertEqual(a.venue, "Kampala Conference Centre")
        self.assertEqual(a.programme_activity_type, "student_activities")
        self.assertEqual(a.programme_delivery_mode, "group")
        self.assertEqual(a.planned_school_count, 12)
        self.assertEqual(a.focus_intervention, "christlike_behaviour")
        self.assertEqual(a.responsible_staff_id, self.cceo.staff_profile_id)
        self.assertIsNone(a.end_date)
        self.assertEqual(a.planned_date, date(2026, 8, 31))

    def test_school_is_not_required_for_non_school_activity(self):
        a = _schedule(self.cceo, districtId=self.district.id)
        self.assertIsNone(a.school_id)
        self.assertIsNone(a.cluster_id)
        self.assertEqual(a.event_district_id, self.district.id)

    def test_project_scoped_activity_carries_programme_context(self):
        # Project-scoped Catalogue work is itself governed: the item must be
        # approved for the Special Project via an active project mapping.
        from apps.activity_catalogue.models import ActivityProjectMapping
        from apps.activity_catalogue.services import get_selectable_item

        ActivityProjectMapping.objects.create(
            project=self.project,
            catalogue_item=get_selectable_item("STUDENT_CONFERENCE_CAMPS"),
            active=True,
        )
        a = _schedule(
            self.cceo,
            supportRationale="project_objective",
            projectId=self.project.id,
        )
        self.assertEqual(a.activity_context_type, "programme")
        self.assertEqual(a.project_id, self.project.id)

    def test_multi_day_activity_is_supported(self):
        a = _schedule(self.cceo, scheduledDate=TUE_SEP_1, endDate="2026-09-03")
        self.assertEqual(a.planned_date, date(2026, 9, 1))
        self.assertEqual(a.end_date, date(2026, 9, 3))
        lines = list(a.schedule_cost_lines.all())
        # One line per component, no month suffixes for a single-month span.
        self.assertEqual(len(lines), 6)
        self.assertFalse(any("#m" in ln.cost_setting_key for ln in lines))
        # 3 days × (venue+facilitation+transport) + 40×3 meals + 40 materials
        # + 2 nights accommodation = 3,660,000
        self.assertEqual(a.est_cost_cents, THREE_DAY_TOTAL)
        self.assertEqual(sum(ln.amount for ln in lines), a.est_cost_cents)

    def test_identical_double_submit_is_refused(self):
        _schedule(self.cceo)
        with self.assertRaisesMessage(BadRequest, "identical activity"):
            psvc.schedule_programme_activity(_payload(), self.cceo)
        self.assertEqual(
            Activity.objects.filter(planning_source="manual_work_plan").count(), 1
        )


# ── Validation ───────────────────────────────────────────────────────────────
class ProgrammeActivityValidationTest(_ProgrammeFixture):
    def test_activity_requires_valid_rationale(self):
        for bad in (_OMIT, "", "because-i-said-so"):
            with self.subTest(rationale=bad):
                with self.assertRaisesMessage(BadRequest, "strategic rationale"):
                    psvc.schedule_programme_activity(
                        _payload(supportRationale=bad), self.cceo
                    )
        self.assertEqual(Activity.objects.count(), 0)

    def test_project_objective_rationale_requires_a_project(self):
        with self.assertRaisesMessage(BadRequest, "Special Project"):
            psvc.schedule_programme_activity(
                _payload(supportRationale="project_objective"), self.cceo
            )

    def test_participant_count_is_required_for_per_head_priced_items(self):
        with self.assertRaisesMessage(BadRequest, "participant count"):
            psvc.schedule_programme_activity(
                _payload(expectedParticipants=_OMIT), self.cceo
            )

    def test_end_date_cannot_precede_start_date(self):
        with self.assertRaisesMessage(BadRequest, "cannot precede"):
            psvc.schedule_programme_activity(
                _payload(scheduledDate=TUE_SEP_1, endDate="2026-08-28"),
                self.cceo,
            )

    def test_programme_activity_span_is_capped_at_31_days(self):
        # Mon 3 Aug → Thu 3 Sep 2026 is a 32-day span.
        with self.assertRaisesMessage(BadRequest, "at most 31 days"):
            psvc.schedule_programme_activity(
                _payload(
                    scheduledDate="2026-08-03T09:00:00+03:00",
                    endDate="2026-09-03",
                ),
                self.cceo,
            )

    def test_multi_day_activity_needs_its_start_date(self):
        with self.assertRaisesMessage(
            BadRequest, "multi-day Activity needs its start date"
        ):
            psvc.schedule_programme_activity(
                _payload(scheduledDate=_OMIT, endDate="2026-09-03"), self.cceo
            )

    def test_school_bound_item_can_also_be_planned_as_a_central_budget_activity(self):
        activity = _schedule(self.cceo, catalogueItemId="CLIENT_SCHOOL_FOLLOWUP_VISIT")

        self.assertIsNone(activity.school_id)
        self.assertEqual(activity.planning_source, "manual_work_plan")
        self.assertEqual(
            activity.catalogue_item.stable_code, "CLIENT_SCHOOL_FOLLOWUP_VISIT"
        )

    def test_non_catalogue_item_is_refused(self):
        with self.assertRaises(BadRequest):
            psvc.schedule_programme_activity(
                _payload(catalogueItemId="NOT_A_REAL_STABLE_CODE"), self.cceo
            )

    def test_sunday_start_date_is_blocked(self):
        with self.assertRaisesMessage(BadRequest, "Sunday"):
            psvc.schedule_programme_activity(
                _payload(scheduledDate=SUN_AUG_30), self.cceo
            )

    def test_sunday_end_date_is_blocked(self):
        # Fri 4 Sep → Sun 6 Sep 2026: the end-date REG-02 check must fire.
        with self.assertRaisesMessage(BadRequest, "Sunday"):
            psvc.schedule_programme_activity(
                _payload(scheduledDate=FRI_SEP_4, endDate="2026-09-06"),
                self.cceo,
            )


# ── Permission gate ──────────────────────────────────────────────────────────
class ProgrammeActivityPermissionTest(_ProgrammeFixture):
    AUTHORIZED_ROLES = (
        "CCEO",
        "Program Lead",
        "ProjectCoordinator",
        "CountryDirector",
        "ImpactAssessment",
        "HumanResources",
        "Admin",
    )

    def test_every_authorized_role_can_schedule(self):
        for idx, role in enumerate(self.AUTHORIZED_ROLES):
            with self.subTest(role=role):
                principal = _person(f"role-{idx}", role)
                a = _schedule(principal)
                self.assertEqual(a.status, "scheduled")

    def test_accountant_cannot_create_programme_activity(self):
        accountant = _person("acct", "Accountant")
        with self.assertRaises(Forbidden):
            psvc.schedule_programme_activity(_payload(), accountant)
        self.assertEqual(Activity.objects.count(), 0)

    def test_partner_roles_cannot_create_programme_activity(self):
        for idx, role in enumerate(("PartnerFieldOfficer", "PartnerAdmin")):
            with self.subTest(role=role):
                partner = _person(f"partner-{idx}", role)
                with self.assertRaises(Forbidden):
                    psvc.schedule_programme_activity(_payload(), partner)
        self.assertEqual(Activity.objects.count(), 0)


# ── Costing (§9) ─────────────────────────────────────────────────────────────
class ProgrammeCostingTest(_ProgrammeFixture):
    def test_cross_month_costs_allocate_correctly(self):
        # Mon 31 Aug → Wed 2 Sep 2026, 40 participants: 1 August day and
        # 2 September days. Every per-day component splits 1:2 by service
        # day, and the split preserves the total exactly.
        a = _schedule(self.cceo, scheduledDate=MON_AUG_31, endDate="2026-09-02")
        lines = list(a.schedule_cost_lines.all())
        total = THREE_DAY_TOTAL
        self.assertEqual(a.est_cost_cents, total)
        self.assertEqual(sum(ln.amount for ln in lines), total)

        by_key = {ln.cost_setting_key: ln.amount for ln in lines}
        self.assertEqual(
            by_key,
            {
                # August hosts 1 of the 3 service days.
                "programme_venue_per_day#m202608": VENUE,
                "programme_participant_meal_cost_per_head#m202608": 40 * MEAL,
                "programme_facilitation_per_day#m202608": FACILITATION,
                "programme_transport_per_day#m202608": TRANSPORT,
                # Materials are a one-off on the first service day.
                "programme_materials_per_participant": 40 * MATERIALS,
                # Two accommodation nights: Aug 31 and Sep 1.
                "programme_accommodation_per_night#m202608": ACCOMMODATION,
                # September hosts the other 2.
                "programme_venue_per_day#m202609": 2 * VENUE,
                "programme_participant_meal_cost_per_head#m202609": 80 * MEAL,
                "programme_facilitation_per_day#m202609": 2 * FACILITATION,
                "programme_transport_per_day#m202609": 2 * TRANSPORT,
                "programme_accommodation_per_night#m202609": ACCOMMODATION,
            },
        )

        august = [ln for ln in lines if ln.month == 8]
        september = [ln for ln in lines if ln.month == 9]
        self.assertEqual(
            sum(ln.amount for ln in august),
            VENUE
            + 40 * MEAL
            + FACILITATION
            + TRANSPORT
            + 40 * MATERIALS
            + ACCOMMODATION,
        )
        self.assertEqual(
            sum(ln.amount for ln in september),
            2 * VENUE + 80 * MEAL + 2 * FACILITATION + 2 * TRANSPORT + ACCOMMODATION,
        )
        # The two months re-sum to the whole: a cross-period activity never
        # costs less than the same activity inside one month.
        self.assertEqual(
            sum(ln.amount for ln in august) + sum(ln.amount for ln in september),
            total,
        )

        # Every line's period stamps follow its OWN service date.
        for ln in lines:
            self.assertEqual(ln.month, ln.planned_date.month)
            self.assertEqual(ln.fiscal_year, FY)
            self.assertEqual(ln.quarter, "Q4")
        for ln in august:
            self.assertEqual(ln.planned_date, date(2026, 8, 31))
        for ln in september:
            self.assertEqual(ln.planned_date, date(2026, 9, 1))

    def test_a_cross_month_split_never_drops_a_component(self):
        """Regression: the first allocator divided each line's QUANTITY by the
        day count, so a component whose quantity does not scale per day
        (venue: qty 1) computed `1 // 3 == 0` and was dropped entirely —
        the activity silently cost less for spanning a month boundary."""
        same_month = _schedule(self.cceo, scheduledDate=TUE_SEP_1, endDate="2026-09-03")
        crossing = _schedule(self.pl, scheduledDate=MON_AUG_31, endDate="2026-09-02")
        self.assertEqual(same_month.est_cost_cents, crossing.est_cost_cents)
        self.assertEqual(
            sum(ln.amount for ln in same_month.schedule_cost_lines.all()),
            sum(ln.amount for ln in crossing.schedule_cost_lines.all()),
        )
        crossing_keys = {
            ln.cost_setting_key.split("#")[0]
            for ln in crossing.schedule_cost_lines.all()
        }
        self.assertEqual(
            crossing_keys,
            {ln.cost_setting_key for ln in same_month.schedule_cost_lines.all()},
        )

    def test_single_month_event_has_no_month_suffixes(self):
        a = _schedule(self.cceo, scheduledDate=TUE_SEP_1, endDate="2026-09-03")
        keys = [ln.cost_setting_key for ln in a.schedule_cost_lines.all()]
        self.assertEqual(len(keys), len(set(keys)))
        self.assertFalse(any("#m" in k for k in keys))
        self.assertEqual(len(keys), 6)

    def test_no_silent_zero_cost(self):
        CostSetting.objects.filter(key="programme_venue_per_day").delete()
        before = Activity.objects.count()
        with self.assertRaises(BadRequest):
            psvc.schedule_programme_activity(_payload(), self.cceo)
        self.assertEqual(Activity.objects.count(), before)
        self.assertEqual(ActivityScheduleCostLine.objects.count(), 0)


class ProgrammeWorkPlanSurfaceTest(_ProgrammeFixture):
    def test_drawer_contains_the_required_planning_fields(self):
        from django.test import Client

        client = Client()
        client.force_login(self.cceo)
        response = client.get("/work-plan/add")
        self.assertEqual(response.status_code, 200)
        html = response.content.decode()
        for field_name in (
            "catalogue_item_id",
            "support_rationale",
            "purpose",
            "focus_intervention",
            "scheduled_date",
            "end_date",
            "planned_school_count",
            "expected_participants",
            "responsibility_type",
            "venue",
        ):
            self.assertIn(f'name="{field_name}"', html)

        # Activity type and delivery mode are DERIVED from the chosen
        # catalogue item, not entered beside it. They used to be their own
        # selects, which let a scheduler pick a type that contradicted the
        # catalogue entry the cost was then drawn from. The catalogue is the
        # governed source, so the drawer publishes them as data attributes on
        # each option and displays them read-only.
        for derived in ('data-activity-type="', 'data-delivery="'):
            self.assertIn(derived, html)
        for retired in ("programme_activity_type", "programme_delivery_mode"):
            self.assertNotIn(f'name="{retired}"', html)

        # Real catalogue entries, grouped by their programme category.
        for label in (
            "School Leadership",
            "Teacher Leadership Conference/Camp/Retreat",
            "Programme",
            "Student Programmes",
        ):
            self.assertIn(label, html)

    def test_excel_export_matches_the_work_plan_columns(self):
        from django.test import Client

        cd = _person("export-cd", "CountryDirector")
        _schedule(cd)
        client = Client()
        client.force_login(cd)
        response = client.get(f"/work-plan/export.xlsx?fy={FY}&view=fy")
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        sheet = workbook["Work Plan"]
        headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        self.assertEqual(
            headers,
            [
                "Activity Date",
                "Activity Title",
                "Activity Type",
                "SSA Intervention",
                "Number of Schools",
                "Number of Participants",
                "Responsible Party",
                "Party Type",
                "Venue",
                "Delivery Mode",
                "Cost (UGX)",
                "Status",
            ],
        )
        data = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        self.assertEqual(data[1], "Student Conference/Camps - Student camps")
        self.assertEqual(data[2], "Student Activities")
        self.assertEqual(data[4], 12)
        self.assertEqual(data[5], 40)
        self.assertEqual(data[9], "Group")

    def test_work_plan_renders_the_premium_matching_columns(self):
        from django.test import Client

        cd = _person("table-cd", "CountryDirector")
        _schedule(cd)
        client = Client()
        client.force_login(cd)
        response = client.get(f"/work-plan?fy={FY}&view=fy")
        self.assertEqual(response.status_code, 200)
        html = response.content.decode()
        for heading in (
            "Activity date",
            "SSA intervention",
            "Scale",
            "Responsible party",
            "Venue &amp; mode",
            "Status &amp; action",
        ):
            self.assertIn(heading, html)
        self.assertIn("Submit completed plan to RVP", html)
        self.assertIn("Export Excel", html)


class ProgrammeWorkPlanApprovalTest(_ProgrammeFixture):
    def test_cd_submits_completed_plan_and_rvp_approves_it(self):
        from apps.monthly_work_plan.models import CountryAnnualBudgetStatus
        from apps.monthly_work_plan.services import rvp_annual_decide
        from apps.planning.work_plan_approval import submit_to_rvp

        cd = _person("approval-cd", "CountryDirector")
        rvp = _person("approval-rvp", "RegionalVicePresident")
        activity = _schedule(cd)

        budget = submit_to_rvp(FY, cd)
        self.assertEqual(budget.status, CountryAnnualBudgetStatus.SUBMITTED_TO_RVP)
        self.assertEqual(budget.target_activities, 1)
        self.assertEqual(budget.target_schools, 12)
        self.assertEqual(budget.program_total, activity.est_cost_cents)

        budget = rvp_annual_decide(budget.id, "approve", {}, rvp)
        self.assertEqual(budget.status, CountryAnnualBudgetStatus.APPROVED_BY_RVP)
        self.assertIsNotNone(budget.baseline_locked_at)


# ── Budget flow, cancel, reschedule ──────────────────────────────────────────
class ProgrammeFundingFlowTest(_ProgrammeFixture):
    def _cross_month_activity(self) -> Activity:
        return _schedule(self.cceo, scheduledDate=MON_AUG_31, endDate="2026-09-02")

    def test_budget_flow_creates_advances_and_weekly_draft(self):
        a = self._cross_month_activity()
        lines = list(a.schedule_cost_lines.all())
        advances = AdvanceRequest.objects.filter(activity=a)
        self.assertEqual(advances.count(), len(lines))
        for adv in advances:
            self.assertEqual(adv.responsible_user_id, self.cceo.id)

        # A weekly draft exists for the week of each line's week_start_date,
        # and each draft's total equals exactly the lines it carries.
        week_starts = {ln.week_start_date for ln in lines}
        for ws in week_starts:
            wfr = WeeklyFundRequest.objects.get(
                responsible_user=self.cceo.id, week_start_date=ws
            )
            self.assertEqual(
                wfr.total_amount,
                sum(wl.total_cost for wl in wfr.lines.all()),
            )
        self.assertEqual(
            WeeklyFundRequestLine.objects.filter(
                activity_budget_line__activity=a
            ).count(),
            len(lines),
        )

    def test_activity_appears_on_responsible_persons_my_plan(self):
        a = _schedule(self.cceo)
        feed = my_plan.get(self.cceo, {"period": "fy", "fy": FY})
        self.assertIn(a.id, {item["id"] for item in feed["items"]})

    def test_assignment_notifies_the_responsible_staff(self):
        a = _schedule(self.pl, responsibleStaffId=self.cceo.staff_profile_id)
        self.assertEqual(a.responsible_staff_id, self.cceo.staff_profile_id)
        self.assertTrue(
            Notification.objects.filter(
                recipient_id=self.cceo.id,
                source_event_type="activity_assigned",
                context_id=a.id,
            ).exists()
        )
        # The assignee sees it on THEIR My Plan; the scheduling PL does not.
        assignee_feed = my_plan.get(self.cceo, {"period": "fy", "fy": FY})
        self.assertIn(a.id, {item["id"] for item in assignee_feed["items"]})

    def test_cancel_withdraws_programme_funding(self):
        a = self._cross_month_activity()
        line_count = a.schedule_cost_lines.count()
        self.assertTrue(AdvanceRequest.objects.filter(activity=a).exists())

        asvc.cancel(a.id, {"reason": "Called off"}, self.cceo)

        a.refresh_from_db()
        self.assertEqual(a.status, "cancelled")
        self.assertFalse(AdvanceRequest.objects.filter(activity=a).exists())
        self.assertEqual(
            WeeklyFundRequestLine.objects.filter(
                activity_budget_line__activity=a
            ).count(),
            0,
        )
        # Cost lines survive as the historical snapshot.
        self.assertEqual(a.schedule_cost_lines.count(), line_count)
        for wfr in WeeklyFundRequest.objects.all():
            self.assertEqual(
                wfr.total_amount,
                sum(wl.total_cost for wl in wfr.lines.all()),
            )

    def test_reschedule_moves_both_dates_and_monthly_attribution(self):
        a = self._cross_month_activity()
        self.assertTrue(a.schedule_cost_lines.filter(month=8).exists())

        asvc.reschedule(
            a.id,
            {"scheduledDate": MON_SEP_14, "reason": "Venue clash"},
            self.cceo,
        )

        a.refresh_from_db()
        # Duration preserved: 3-day event moves start AND end together.
        self.assertEqual(a.planned_date, date(2026, 9, 14))
        self.assertEqual(a.end_date, date(2026, 9, 16))
        self.assertEqual(a.status, "rescheduled")

        lines = list(a.schedule_cost_lines.all())
        # Re-priced as a single-month event: 6 unsuffixed component lines,
        # all attributed to September, same total.
        self.assertEqual(len(lines), 6)
        self.assertFalse(any("#m" in ln.cost_setting_key for ln in lines))
        self.assertTrue(all(ln.month == 9 for ln in lines))
        self.assertTrue(all(ln.planned_date == date(2026, 9, 14) for ln in lines))
        self.assertEqual(a.est_cost_cents, THREE_DAY_TOTAL)
        self.assertEqual(sum(ln.amount for ln in lines), THREE_DAY_TOTAL)

        # The vacated August week's draft is gone; the new week carries all.
        self.assertFalse(
            WeeklyFundRequest.objects.filter(week_start_date=date(2026, 8, 31)).exists()
        )
        new_wfr = WeeklyFundRequest.objects.get(
            responsible_user=self.cceo.id,
            week_start_date=date(2026, 9, 14),
        )
        self.assertEqual(new_wfr.total_amount, THREE_DAY_TOTAL)


# ── Health (§35) ─────────────────────────────────────────────────────────────
class ProgrammeWorkPlanHealthTest(_ProgrammeFixture):
    def test_work_plan_health_passes_after_wellformed_schedule(self):
        _schedule(self.cceo, scheduledDate=MON_AUG_31, endDate="2026-09-02")
        health = work_plan_health()
        self.assertTrue(health["healthy"])
        for check in health["checks"]:
            self.assertNotEqual(check["status"], "fail")

    def test_health_flags_programme_activity_without_rationale(self):
        # A raw ORM write that bypasses the service funnel must be caught.
        Activity.objects.create(
            activity_type="training",
            planning_source="manual_work_plan",
            delivery_type="staff",
            status="scheduled",
            fy=FY,
            quarter="Q4",
            planned_date=date(2026, 9, 1),
            responsible_staff_id=self.cceo.staff_profile_id,
            support_rationale="",
        )
        health = work_plan_health()
        check = next(
            c
            for c in health["checks"]
            if c["key"] == "programme_activity_without_rationale"
        )
        self.assertEqual(check["status"], "fail")
        self.assertEqual(check["count"], 1)
        self.assertFalse(health["healthy"])


# ── Concurrency (TransactionTestCase, mirrors test_audit_pipeline) ───────────
class ProgrammeDoubleClickRaceTest(TransactionTestCase):
    """Two identical concurrent submissions must yield exactly ONE activity.

    Non-school work has no school/cluster row to lock; activities.create
    serialises on the responsible StaffProfile row instead, so the loser's
    duplicate check runs after the winner's commit."""

    reset_sequences = False

    def setUp(self):
        self.cceo = _person("race-cceo", "CCEO")

    def test_double_click_create_creates_one_activity(self):
        results = [None, None]
        barrier = threading.Barrier(2)

        def submit(idx):
            try:
                barrier.wait(timeout=10)
                results[idx] = (
                    True,
                    psvc.schedule_programme_activity(_payload(), self.cceo),
                )
            except Exception as exc:  # noqa: BLE001 - collected for assertion
                results[idx] = (False, exc)
            finally:
                for conn in connections.all():
                    conn.close()

        threads = [threading.Thread(target=submit, args=(i,)) for i in (0, 1)]
        for t in threads:
            t.start()
        for t in threads:
            t.join(timeout=30)

        winners = [r for r in results if r and r[0]]
        losers = [r for r in results if r and not r[0]]
        self.assertEqual(len(winners), 1, f"results: {results}")
        self.assertEqual(len(losers), 1, f"results: {results}")
        self.assertIsInstance(losers[0][1], BadRequest)
        self.assertIn("identical activity", str(losers[0][1]))
        self.assertEqual(
            Activity.objects.filter(planning_source="manual_work_plan").count(),
            1,
        )


# ── §19 My Plan handoff ──────────────────────────────────────────────────────
class ProgrammeMyPlanHandoffTest(_ProgrammeFixture):
    """A school-less programme activity must reach the responsible person's
    My Plan — and must not take the page down on the way.

    Regression: every school/cluster attribute in the My Plan row builder
    assumed one of the two existed. `a.school.cluster_id` ran whenever there
    was no cluster, so the FIRST programme activity assigned to anyone 500'd
    their entire My Plan page — the feature's own §19 handoff was also its
    biggest outage.
    """

    def test_my_plan_service_includes_the_programme_activity(self):
        from apps.my_plan.services import get_frontend_context

        activity = _schedule(self.pl)
        ctx = get_frontend_context(
            self.pl,
            {
                "period": "month",
                "fy": activity.fy,
                "month": str(activity.planned_month),
            },
        )
        ids = {row["id"] for row in ctx["programme_activities_all"]}
        self.assertIn(activity.id, ids)

    def test_my_plan_page_renders_without_a_school(self):
        from django.test import Client

        activity = _schedule(self.pl)
        client = Client()
        client.force_login(self.pl)
        response = client.get(
            f"/my-plan?period=month&fy={activity.fy}&month={activity.planned_month}"
        )
        self.assertEqual(response.status_code, 200)
        html = response.content.decode()
        self.assertIn("Programme Activities Planned", html)
        self.assertIn(activity.id, html)

    def test_programme_row_names_the_venue_not_unknown_school(self):
        from apps.my_plan.services import get_frontend_context

        activity = _schedule(self.pl, venue="Kampala Serena")
        ctx = get_frontend_context(
            self.pl,
            {
                "period": "month",
                "fy": activity.fy,
                "month": str(activity.planned_month),
            },
        )
        row = next(r for r in ctx["programme_activities_all"] if r["id"] == activity.id)
        self.assertEqual(row["school_name"], "Kampala Serena")
        self.assertNotEqual(row["school_name"], "Unknown School")
