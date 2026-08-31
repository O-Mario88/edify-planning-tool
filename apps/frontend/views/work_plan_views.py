"""Work Plan: the Add Non-School Activity entry point (§5-§8).

One drawer, one backend service. Both the Fund/Budget dashboard and the Work
Plan page open THIS drawer, and its submit calls
apps.planning.services.schedule_programme_activity — the canonical Activity
funnel. Nothing here computes cost or invents approval routing.
"""

from __future__ import annotations

from functools import wraps
from urllib.parse import quote

from django.contrib import messages
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import redirect, render
from django.views.decorators.http import require_http_methods

from apps.core.enums import (
    ProgrammeActivityType,
    ProgrammeDeliveryMode,
    SsaIntervention,
    SupportRationale,
)
from apps.core.htmx_errors import error_fragment
from apps.core.permissions import has_permission
from apps.core.rbac import Permission
from apps.accounts.staff_matching import on_staff


def _signed_out_redirect(request):
    """Answer a signed-out visitor the way the rest of the product does.

    Every other guarded route redirects to /login preserving the destination;
    these Work Plan routes answered a bare 403, so a logged-out person
    following a link lost where they were going and got an error page instead
    of a sign-in form.

    Only safe methods redirect. A signed-out POST is a mutation attempt, and
    403 is the honest machine-readable answer there — bouncing it to a login
    page would drop the body and imply the request could be replayed.
    Authenticated-but-unauthorized always stays 403: that is a real refusal.
    """
    if request.method not in ("GET", "HEAD"):
        return HttpResponseForbidden("Sign in first.")

    from django.contrib.auth.views import redirect_to_login

    return redirect_to_login(request.get_full_path(), login_url="/login")


def _manual_activity_permission(view):
    @wraps(view)
    def wrapped(request, *args, **kwargs):
        if not getattr(request.user, "is_authenticated", False):
            return _signed_out_redirect(request)
        if not has_permission(request.user, Permission.MANUAL_ACTIVITY_CREATE.value):
            return HttpResponseForbidden(
                "You do not have permission to add non-school programme " "activities."
            )
        return view(request, *args, **kwargs)

    # The production-readiness route scanner recognises guarded page views
    # by this attribute (the same contract require_page_permission uses).
    wrapped.has_permission_guard = True
    return wrapped


def _programme_items():
    from apps.activity_catalogue.services import effective_items

    return (
        effective_items()
        .filter(non_school_allowed=True)
        .prefetch_related("intervention_mappings")
        .order_by("programme_category", "display_name")
    )


def _programme_item_metadata(item):
    """Canonical reporting fields for a central budget activity.

    The catalogue owns activity type, delivery method and fixed intervention;
    asking planners to select them again created contradictory combinations.
    """
    from apps.activity_catalogue.models import DeliveryMethod, MappingMode
    from apps.core.enums import ProgrammeDeliveryMode

    delivery_modes = {
        DeliveryMethod.IN_SCHOOL_TRAINING: ProgrammeDeliveryMode.IN_SCHOOL,
        DeliveryMethod.CLUSTER_MEETING: ProgrammeDeliveryMode.CLUSTER,
        DeliveryMethod.CLUSTER_TRAINING: ProgrammeDeliveryMode.CLUSTER,
        DeliveryMethod.ONLINE: ProgrammeDeliveryMode.ONLINE,
        DeliveryMethod.SCHOOL_VISIT: ProgrammeDeliveryMode.VISIT,
        DeliveryMethod.GROUP: ProgrammeDeliveryMode.GROUP,
        DeliveryMethod.ADMIN: ProgrammeDeliveryMode.ADMIN,
        DeliveryMethod.PROGRAMME_EVENT: ProgrammeDeliveryMode.GROUP,
    }
    mapping = next(
        (
            mapping
            for mapping in item.intervention_mappings.all()
            if mapping.active and mapping.is_primary
        ),
        None,
    )
    requires_intervention = bool(
        mapping
        and mapping.mapping_mode == MappingMode.INHERIT_FROM_SOURCE_ACTIVITY
        and not mapping.intervention
    )
    return {
        "programme_activity_type": item.activity_type,
        "programme_activity_type_label": item.get_activity_type_display(),
        "programme_delivery_mode": delivery_modes[item.delivery_method],
        "programme_delivery_mode_label": item.get_delivery_method_display(),
        "focus_intervention": mapping.intervention if mapping else None,
        "requires_intervention": requires_intervention,
        "focus_intervention_label": (
            mapping.get_intervention_display()
            if mapping and mapping.intervention
            else (
                "Select follow-up intervention"
                if requires_intervention
                else "Administrative / not applicable"
            )
        ),
        "is_training_course": item.is_training_course,
        "training_category": item.training_category,
        "ssa_indicator_label": item.ssa_indicator_label,
    }


def _assignable_staff(user):
    """Who may this user make responsible for a programme activity?

    Everyone may plan their own work; supervisory roles (PL, CD, HR, Admin)
    may assign another active staff member. Mirrors §6C.
    """
    from apps.accounts.models import StaffProfile

    if user.active_role not in (
        "Program Lead",
        "CountryDirector",
        "HumanResources",
        "Admin",
    ):
        return StaffProfile.objects.none()
    return (
        # on_staff, not is_active: a pending-invite CCEO created by a school upload already holds their portfolio.
        on_staff(StaffProfile.objects).select_related("user").order_by("user__name")
    )


def _active_partners():
    from apps.partners.models import Partner

    return Partner.objects.filter(
        deleted_at__isnull=True,
        active_status=True,
    ).order_by("name")


@_manual_activity_permission
def non_school_activity_drawer(request):
    from apps.geography.models import District
    from apps.projects.scoping import scoped_projects

    items = list(_programme_items())
    for item in items:
        item.programme_metadata = _programme_item_metadata(item)

    date_prefill = (request.GET.get("date") or "").strip()[:10]
    return render(
        request,
        "partials/work_plan/non_school_activity_drawer.html",
        {
            "date_prefill": date_prefill,
            "items": items,
            "rationales": SupportRationale.choices,
            "projects": scoped_projects(request.user).order_by("name"),
            "districts": District.objects.order_by("name"),
            "assignable_staff": _assignable_staff(request.user),
            "partners": _active_partners(),
            "programme_activity_types": ProgrammeActivityType.choices,
            "programme_delivery_modes": ProgrammeDeliveryMode.choices,
            "ssa_interventions": SsaIntervention.choices,
            "drawer_type": "right",
            "drawer_size": "lg",
        },
    )


def _travel_district_type(dest_district_id: str, responsible_id: str) -> str:
    """Preview-side twin of activities.services._field_event_district_type."""
    if not dest_district_id:
        return "primary"
    from django.db.models import Q

    from apps.accounts.models import StaffProfile
    from apps.geography.models import District

    home = (
        StaffProfile.objects.filter(Q(id=responsible_id) | Q(user_id=responsible_id))
        .values_list("primary_district_id", flat=True)
        .first()
    )
    if home:
        return "primary" if home == dest_district_id else "secondary"
    dtype = (
        District.objects.filter(id=dest_district_id)
        .values_list("district_type", flat=True)
        .first()
    )
    return dtype or "secondary"


@_manual_activity_permission
@require_http_methods(["POST"])
def non_school_activity_preview(request):
    """Live catalogue cost preview for the drawer — read-only, no writes."""
    from datetime import date

    from apps.budget.costing_service import preview as cost_preview

    try:
        start_raw = (request.POST.get("scheduled_date") or "").strip()
        end_raw = (request.POST.get("end_date") or "").strip()
        days = 1
        if start_raw and end_raw:
            start = date.fromisoformat(start_raw[:10])
            end = date.fromisoformat(end_raw[:10])
            if end >= start:
                days = (end - start).days + 1
        participants = int(request.POST.get("expected_participants") or 0) or None
        from apps.activity_catalogue.services import get_selectable_item

        item_id = (request.POST.get("catalogue_item_id") or "").strip()
        if not item_id:
            raise ValueError("Choose an Activity title to preview its cost.")
        item = get_selectable_item(item_id)
        payload = {
            "activityType": item.workflow_kind,
            "costingProfile": item.costing_profile,
            "deliveryType": (
                "partner"
                if request.POST.get("responsibility_type") == "partner"
                else "staff"
            ),
            "expectedParticipants": participants,
            "days": days,
        }
        travel_profile = None
        if item.costing_profile == "FIELD_TRAVEL":
            responsible = (
                (request.POST.get("responsible_staff_id") or "").strip()
                or request.user.staff_profile_id
                or request.user.user_id
            )
            district_type = _travel_district_type(
                (request.POST.get("district_id") or "").strip(), responsible
            )
            payload["districtType"] = district_type
            if district_type == "secondary":
                # Every secondary away-day carries the full per-diem set —
                # meals and a night's accommodation per day (owner rule,
                # 2026-08-19), matching the secondary visit-day policy.
                travel_profile = (
                    f"Another district · {days} day"
                    f"{'' if days == 1 else 's'} · full per-diems per day"
                )
            else:
                travel_profile = "Home district — transport and lunch per day"
        result = cost_preview(payload, minimum=True)
    except Exception as exc:  # noqa: BLE001 — preview must degrade, not 500
        return error_fragment(exc, action="Could not preview the cost", status=400)
    return render(
        request,
        "partials/work_plan/non_school_cost_preview.html",
        {"preview": result, "days": days, "travel_profile": travel_profile},
    )


def _district_name(district_id) -> str:
    from apps.geography.models import District

    district_id = (district_id or "").strip()
    if not district_id:
        return ""
    return (
        District.objects.filter(id=district_id).values_list("name", flat=True).first()
        or ""
    )


@_manual_activity_permission
@require_http_methods(["POST"])
def non_school_activity_action(request):
    from apps.planning.services import schedule_programme_activity
    from apps.activity_catalogue.services import get_selectable_item

    # Simplified drawer (2026-08-19): the planner schedules their OWN
    # attendance, and the governed catalogue item is the authorization. The
    # removed inputs default here so the create funnel still receives a
    # complete, honest payload. Legacy callers that still post the old
    # fields keep working — an explicit value always wins over a default.
    responsibility_type = (request.POST.get("responsibility_type") or "staff").strip()
    if responsibility_type not in ("staff", "partner"):
        return error_fragment(
            ValueError("Select Staff or Partner as the responsible party."),
            action="Could not add the activity",
            status=400,
        )

    responsible = (request.POST.get("responsible_staff_id") or "").strip()
    partner_id = (request.POST.get("partner_id") or "").strip()
    own_id = request.user.staff_profile_id or request.user.user_id
    if responsibility_type == "staff":
        responsible = responsible or own_id
        if (
            responsible != own_id
            and not _assignable_staff(request.user).filter(id=responsible).exists()
        ):
            return HttpResponseForbidden(
                "Only an authorized supervisor may assign this activity to "
                "another staff member."
            )
    elif not partner_id or not _active_partners().filter(id=partner_id).exists():
        return error_fragment(
            ValueError("Select an active Partner responsible for the activity."),
            action="Could not add the activity",
            status=400,
        )

    try:
        participants = int(request.POST.get("expected_participants") or 0)
        schools = int(request.POST.get("planned_school_count") or 0)
    except (TypeError, ValueError) as exc:
        return error_fragment(
            exc,
            action="Schools and participants must be whole numbers",
            status=400,
        )

    start_raw = (request.POST.get("scheduled_date") or "").strip()
    catalogue_item_id = (request.POST.get("catalogue_item_id") or "").strip()
    try:
        item = get_selectable_item(catalogue_item_id)
        metadata = _programme_item_metadata(item)
    except Exception as exc:  # noqa: BLE001 — htmx error surface
        return error_fragment(
            exc, action="Choose an approved Activity title", status=400
        )
    payload = {
        "catalogueItemId": catalogue_item_id,
        "scheduledDate": f"{start_raw}T09:00:00+03:00" if start_raw else "",
        "endDate": (request.POST.get("end_date") or "").strip() or None,
        "supportRationale": (
            (request.POST.get("support_rationale") or "").strip()
            or "organizational_priority"
        ),
        "projectId": (request.POST.get("project_id") or "").strip() or None,
        "activityPurposeText": (
            (request.POST.get("purpose") or "").strip() or item.display_name
        ),
        "programmeActivityType": metadata["programme_activity_type"],
        "programmeDeliveryMode": metadata["programme_delivery_mode"],
        "focusIntervention": (
            (request.POST.get("focus_intervention") or "").strip()
            if metadata["requires_intervention"]
            else metadata["focus_intervention"]
        ),
        "plannedSchoolCount": schools,
        "expectedOutcome": (request.POST.get("expected_outcome") or "").strip(),
        "expectedParticipants": participants or None,
        # Venue is optional in the simplified drawer; the funnel requires
        # one, so an unnamed venue defaults to the destination district (or
        # a neutral placeholder) rather than blocking the plan.
        "venue": (
            (request.POST.get("venue") or "").strip()
            or _district_name(request.POST.get("district_id"))
            or "To be confirmed"
        ),
        "districtId": (request.POST.get("district_id") or "").strip() or None,
        "deliveryType": responsibility_type,
    }
    if responsibility_type == "staff":
        payload["responsibleStaffId"] = responsible
    else:
        payload["assignedPartnerId"] = partner_id
    try:
        result = schedule_programme_activity(payload, request.user)
    except Exception as exc:  # noqa: BLE001 — htmx error surface
        return error_fragment(exc, action="Could not add the activity", status=400)

    response = HttpResponse(
        '<div class="p-3 rounded-surface bg-emerald-50 text-emerald-800 '
        'text-[13px] font-bold" role="status">Activity added to the Work Plan.'
        "</div>"
        "<script>window.setTimeout(function () {"
        'window.location.href = "/work-plan";'
        "}, 600);</script>"
    )
    response["HX-Trigger"] = "close-drawer"
    response["X-Activity-Id"] = result["id"]
    return response


def _require_permission(permission: str, denied_message: str):
    def decorator(view):
        @wraps(view)
        def wrapped(request, *args, **kwargs):
            if not getattr(request.user, "is_authenticated", False):
                return _signed_out_redirect(request)
            if not has_permission(request.user, permission):
                return HttpResponseForbidden(denied_message)
            return view(request, *args, **kwargs)

        wrapped.has_permission_guard = True
        return wrapped

    return decorator


def _require_export(view):
    """Compatibility alias kept for route scanners and focused tests."""

    @wraps(view)
    def wrapped(request, *args, **kwargs):
        if not getattr(request.user, "is_authenticated", False):
            return _signed_out_redirect(request)
        if not has_permission(request.user, Permission.EXPORT.value):
            return HttpResponseForbidden("You do not have permission to export.")
        return view(request, *args, **kwargs)

    wrapped.has_permission_guard = True
    return wrapped


@_require_export
@require_http_methods(["GET"])
def work_plan_export(request):
    """Export the same scoped rows and visible fields as the Work Plan table."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    from apps.frontend.views.work_plan_page import build_work_plan_context

    context = build_work_plan_context(request.user, request.GET)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Work Plan"
    headers = [
        "Group",
        "Activity Date",
        "Activity Title",
        "Activity Type",
        "SSA Intervention",
        "Number of Schools",
        "Number of Participants",
        "Responsible Party",
        "Party Type",
        "Venue",
        "Delivery Mode",
        "Cost (UGX)",
        "Status",
    ]
    sheet.append(headers)
    for row in context["rows"]:
        sheet.append(
            [
                row["group_label"],
                row["date_label"],
                row["name"],
                row["programme_activity_type"],
                row["focus_intervention"],
                row["school_count"] or 0,
                row["participants"] or 0,
                row["responsible"],
                row["responsibility_type"],
                row["venue"],
                row["programme_delivery_mode"],
                int(row["cost"] or 0),
                row["status_label"],
            ]
        )

    navy = "102A43"
    accent = "1677FF"
    border = Side(style="thin", color="D9E2EC")
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(vertical="center")
        cell.border = Border(bottom=Side(style="medium", color=accent))
    sheet.row_dimensions[1].height = 28
    for row_index in range(2, sheet.max_row + 1):
        fill = PatternFill(
            "solid", fgColor="F7FAFC" if row_index % 2 == 0 else "FFFFFF"
        )
        for cell in sheet[row_index]:
            cell.fill = fill
            cell.border = Border(bottom=border)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.cell(row_index, 12).number_format = "#,##0"
    widths = [22, 20, 38, 28, 26, 18, 21, 28, 14, 28, 18, 18, 20]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False

    # Summary sheet — the CD/RVP roll-up: one line per (executor, activity
    # type), grouped School vs Non-School. Same aggregates the page renders.
    summary_sheet = workbook.create_sheet("Plan Summary")
    summary_sheet.append(
        ["Activity", "Number of Activities", "Unit Cost (UGX)", "Total Cost (UGX)"]
    )
    for cell in summary_sheet[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(vertical="center")
        cell.border = Border(bottom=Side(style="medium", color=accent))
    summary_sheet.row_dimensions[1].height = 28
    for section in context["plan_summary_sections"]:
        header_row = summary_sheet.max_row + 1
        summary_sheet.append([section["label"], "", "", ""])
        summary_sheet.cell(header_row, 1).font = Font(bold=True, size=10)
        for row in section["rows"]:
            summary_sheet.append(
                [
                    row["label"],
                    row["count"],
                    row["unit_cost"]
                    if row["unit_cost"] is not None
                    else "Cost setup required",
                    row["cost"],
                ]
            )
        subtotal_row = summary_sheet.max_row + 1
        summary_sheet.append(
            [f"{section['label']} subtotal", section["count"], "", section["cost"]]
        )
        for column in (1, 2, 4):
            summary_sheet.cell(subtotal_row, column).font = Font(bold=True, size=10)
    grand_row = summary_sheet.max_row + 1
    summary_sheet.append(
        [
            "Grand total",
            context["plan_summary"]["total"]["count"],
            "",
            context["plan_summary"]["total"]["cost"],
        ]
    )
    for column in (1, 2, 4):
        summary_sheet.cell(grand_row, column).font = Font(bold=True, size=10)
    for row_index in range(2, summary_sheet.max_row + 1):
        for column in (2, 3, 4):
            summary_sheet.cell(row_index, column).number_format = (
                "#,##0.##" if column == 3 else "#,##0"
            )
    summary_sheet.append([])
    summary_sheet.append(["Scope", context["scope_label"]])
    summary_sheet.append(["Period", context["period_label"]])
    summary_sheet.append(
        ["Unit cost", "Average recorded cost per activity; rounded where needed."]
    )
    if context["plan_summary"]["total"]["cost_missing_count"]:
        summary_sheet.append(
            [
                "Cost setup required",
                context["plan_summary"]["total"]["cost_missing_count"],
                "Totals include recorded costs only; no rates estimated.",
            ]
        )
    for index, width in enumerate((36, 22, 20, 22), start=1):
        summary_sheet.column_dimensions[get_column_letter(index)].width = width
    summary_sheet.freeze_panes = "A2"
    summary_sheet.sheet_view.showGridLines = False

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    )
    filename = f"Edify_Work_Plan_FY_{context['fy']}.xlsx"
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
    workbook.save(response)
    return response


@_require_permission(
    Permission.COUNTRY_BUDGET_SUBMIT.value,
    "Only the Country Director can submit the Work Plan.",
)
@require_http_methods(["POST"])
def work_plan_submit(request):
    from apps.core.exceptions import BadRequest, Forbidden
    from apps.planning.work_plan_approval import submit_to_rvp

    fy = (request.POST.get("fy") or "").strip()
    try:
        submit_to_rvp(fy, request.user)
        messages.success(
            request,
            f"FY {fy} Work Plan submitted to the RVP for approval.",
        )
    except (BadRequest, Forbidden) as exc:
        messages.error(request, str(exc))
    return redirect(f"/work-plan?fy={quote(fy)}&view=fy")


@_require_permission(
    Permission.COUNTRY_BUDGET_APPROVE.value,
    "Only the RVP can decide this Work Plan.",
)
@require_http_methods(["POST"])
def work_plan_rvp_decision(request):
    from apps.core.exceptions import BadRequest, Forbidden
    from apps.core.rbac import EdifyRole
    from apps.monthly_work_plan.services import rvp_annual_decide

    if request.user.active_role not in (
        EdifyRole.REGIONAL_VICE_PRESIDENT.value,
        EdifyRole.ADMIN.value,
    ):
        return HttpResponseForbidden("Only the RVP can decide this Work Plan.")
    fy = (request.POST.get("fy") or "").strip()
    budget_id = (request.POST.get("budget_id") or "").strip()
    action = (request.POST.get("action") or "").strip()
    try:
        rvp_annual_decide(
            budget_id,
            action,
            {"note": request.POST.get("note")},
            request.user,
        )
        messages.success(
            request,
            (
                f"FY {fy} Work Plan approved and baseline locked."
                if action == "approve"
                else f"FY {fy} Work Plan returned to the Country Director."
            ),
        )
    except (BadRequest, Forbidden) as exc:
        messages.error(request, str(exc))
    return redirect(f"/work-plan?fy={quote(fy)}&view=fy")
