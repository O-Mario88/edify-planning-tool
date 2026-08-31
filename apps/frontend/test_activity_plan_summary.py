"""Summary/detail parity, genuine planned costs, and work-plan role boundaries."""

from datetime import date
from decimal import Decimal
from io import BytesIO

from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook

from apps.accounts.models import (
    StaffGeographyAssignment,
    StaffProfile,
    StaffSupervisorAssignment,
    User,
)
from apps.activities.models import Activity, ActivityScheduleCostLine
from apps.frontend.views.work_plan_page import build_work_plan_context
from apps.geography.models import District, Region
from apps.schools.models import School


class ActivityPlanSummaryTest(TestCase):
    @classmethod
    def setUpTestData(cls):
        for key, role in (
            ("cceo", "CCEO"),
            ("pl", "Program Lead"),
            ("other", "CCEO"),
            ("cd", "CountryDirector"),
            ("rvp", "RegionalVicePresident"),
        ):
            person = User.objects.create_user(
                email=f"{key}@plan-summary.test",
                name=key,
                roles=[role],
                active_role=role,
            )
            StaffProfile.objects.create(user=person, title=role)
            setattr(cls, key, person)
        StaffSupervisorAssignment.objects.create(
            supervisor_id=cls.pl.staff_profile_id,
            supervisee_id=cls.cceo.staff_profile_id,
        )
        cls.region = Region.objects.create(name="Plan region")
        cls.district = District.objects.create(name="Plan district", region=cls.region)
        cls.school = School.objects.create(
            school_id="PLAN-SUMMARY",
            name="Plan school",
            region=cls.region,
            district=cls.district,
        )
        cls.other_region = Region.objects.create(name="Other plan region")
        cls.other_district = District.objects.create(
            name="Other plan district", region=cls.other_region
        )
        StaffGeographyAssignment.objects.create(
            staff_id=cls.rvp.staff_profile_id, region_id=cls.region.id
        )

    def activity(self, owner=None, amount=75_000, **fields):
        owner = owner or self.cceo
        values = {
            "activity_type": "school_visit",
            "delivery_type": "staff",
            "status": "scheduled",
            "fy": "2026",
            "planned_date": date(2026, 8, 12),
            "responsible_staff_id": owner.id,
            "event_district": self.district,
        }
        values.update(fields)
        activity = Activity.objects.create(**values)
        if amount is not None:
            self.line(activity, amount, activity.planned_date)
        return activity

    def line(self, activity, amount, day, key="transport"):
        return ActivityScheduleCostLine.objects.create(
            activity=activity,
            label="Planned cost",
            cost_setting_key=f"plan_summary_fixture_{key}",
            line_item_type="transport",
            quantity=1,
            unit_cost=amount,
            amount=amount,
            planned_date=day,
            month=day.month,
            fiscal_year="2026",
        )

    def context(self, user=None, **params):
        return build_work_plan_context(
            user or self.cd, {"fy": "2026", "view": "fy", **params}
        )

    def test_summary_counts_activities_once_and_sums_real_cost_lines(self):
        first = self.activity(amount=50_000)
        self.line(first, 25_000, first.planned_date, key="meals")
        self.activity(amount=75_001)
        self.activity(amount=50_000, delivery_type="partner")
        summary = self.context()["plan_summary"]
        by_label = {r["label"]: r for r in summary["school"]["rows"]}
        staff = by_label["Staff School Visit"]
        self.assertEqual((staff["count"], staff["cost"]), (2, 150_001))
        self.assertEqual(staff["unit_cost"], Decimal("75000.50"))
        self.assertEqual(staff["unit_cost_display"], "75,000.50")
        self.assertEqual(by_label["Partner School Visit"]["cost"], 50_000)
        self.assertEqual(summary["total"]["cost"], 200_001)

    def test_school_types_and_named_non_school_events_share_detail_grouping(self):
        for kind in ("school_visit", "training", "cluster_training", "cluster_meeting"):
            self.activity(activity_type=kind)
        for kind, name in (
            ("programme_event", "Conference"),
            ("field_event", "Boot camp"),
            ("field_event", "District meeting"),
        ):
            self.activity(
                activity_type=kind, activity_name_snapshot=name, school=self.school
            )
        context = self.context()
        self.assertEqual(context["plan_summary"]["school"]["count"], 4)
        self.assertEqual(
            {r["label"] for r in context["plan_summary"]["non_school"]["rows"]},
            {"Staff Conference", "Staff Boot camp", "Staff District meeting"},
        )
        self.assertEqual(
            [r["group"] for r in context["rows"]], ["school"] * 4 + ["non_school"] * 3
        )

    def test_program_lead_own_and_team_include_both_owner_ids_and_partner_monitors(
        self,
    ):
        own = [
            self.activity(owner=self.pl),
            self.activity(responsible_staff_id=self.pl.staff_profile_id),
        ]
        team = [
            self.activity(),
            self.activity(responsible_staff_id=self.cceo.staff_profile_id),
        ]
        for monitor in (self.cceo.id, self.cceo.staff_profile_id):
            team.append(
                self.activity(
                    owner=self.other,
                    delivery_type="partner",
                    monitored_by_staff_id=monitor,
                )
            )
        self.activity(owner=self.other)
        own_context = self.context(self.pl, scope="own")
        team_context = self.context(self.pl, scope="team")
        self.assertEqual({r["id"] for r in own_context["rows"]}, {a.id for a in own})
        self.assertEqual(
            {r["id"] for r in team_context["rows"]}, {a.id for a in own + team}
        )
        self.assertEqual(team_context["plan_summary"]["total"]["count"], 6)
        self.assertEqual(self.context(self.pl)["plan_scope"], "team")

    def test_cceo_cannot_expand_scope_or_select_another_persons_plan(self):
        own = self.activity()
        self.activity(owner=self.other)
        context = self.context(self.cceo, scope="team")
        self.assertEqual([r["id"] for r in context["rows"]], [own.id])
        self.assertEqual(context["scope_tabs"], [])
        self.assertEqual(
            self.context(self.cceo, scope="team", responsible=self.other.id)["rows"], []
        )

    def test_non_school_catalogue_training_keeps_its_planned_event_name(self):
        self.activity(
            activity_type="cluster_training",
            planning_source="manual_work_plan",
            activity_name_snapshot="Student Conference and Camps",
        )
        context = self.context()
        self.assertEqual(context["plan_summary"]["school"]["count"], 0)
        self.assertEqual(
            context["plan_summary"]["non_school"]["rows"][0]["label"],
            "Staff Student Conference and Camps",
        )
        self.assertEqual(context["rows"][0]["group"], "non_school")

    def test_undated_legacy_cost_is_not_repeated_in_spillover_month(self):
        activity = self.activity(
            planned_date=date(2026, 8, 31), end_date=date(2026, 9, 2)
        )
        activity.schedule_cost_lines.update(planned_date=None, month=None)
        self.assertEqual(
            self.context(view="month", period="8")["plan_summary"]["total"]["cost"],
            75_000,
        )
        self.assertEqual(
            self.context(view="month", period="9")["plan_summary"]["total"]["cost"], 0
        )

    def test_cd_country_and_rvp_assigned_region_apply_to_summary_and_details(self):
        own_region = self.activity(
            activity_type="field_event", activity_name_snapshot="District meeting"
        )
        self.activity(owner=self.other, event_district=self.other_district)
        self.assertEqual(self.context(self.cd)["plan_summary"]["total"]["count"], 2)
        regional = self.context(self.rvp)
        self.assertEqual([r["id"] for r in regional["rows"]], [own_region.id])
        self.assertEqual(regional["plan_summary"]["total"]["count"], 1)

    def test_cancelled_deleted_and_out_of_period_activities_do_not_pad_totals(self):
        self.activity(amount=120)
        self.activity(status="cancelled")
        self.activity(deleted_at=timezone.now())
        self.activity(planned_date=date(2026, 7, 12))
        context = self.context(view="month", period="8")
        self.assertEqual(context["plan_summary"]["total"]["count"], 1)
        self.assertEqual(context["plan_summary"]["total"]["cost"], 120)

    def test_multi_month_costs_are_not_repeated_in_each_month(self):
        activity = self.activity(
            amount=50_000, planned_date=date(2026, 8, 31), end_date=date(2026, 9, 2)
        )
        self.line(activity, 70_000, date(2026, 9, 1), key="meals")
        for month, total in (("8", 50_000), ("9", 70_000)):
            context = self.context(view="month", period=month)
            self.assertEqual(context["plan_summary"]["total"]["cost"], total)
            self.assertEqual(context["rows"][0]["cost"], total)
            self.assertEqual(sum(g["cost"] for g in context["groups"]), total)
        self.assertEqual(self.context()["plan_summary"]["total"]["cost"], 120_000)

    def test_missing_costs_and_drilldowns_match_summary_and_details(self):
        missing = self.activity(amount=None)
        self.activity(amount=10_000)
        context = self.context(flag="cost_missing")
        self.assertEqual([r["id"] for r in context["rows"]], [missing.id])
        summary = context["plan_summary"]
        self.assertEqual(summary["total"]["count"], 1)
        self.assertEqual(summary["total"]["cost_missing_count"], 1)
        self.assertIsNone(summary["school"]["rows"][0]["unit_cost"])

    def test_detailed_scale_uses_planned_participants_not_actual_attendance(self):
        self.activity(
            expected_participants=20, teachers_attended=50, leaders_attended=10
        )
        self.assertEqual(self.context()["rows"][0]["participants"], 20)

    def test_summary_updates_when_the_source_cost_changes(self):
        activity = self.activity(amount=75_000)
        self.assertEqual(self.context()["plan_summary"]["total"]["cost"], 75_000)
        activity.schedule_cost_lines.update(amount=90_000)
        self.assertEqual(self.context()["plan_summary"]["total"]["cost"], 90_000)

    def test_page_links_summary_to_grouped_details_and_preserves_pl_scope(self):
        self.activity(owner=self.pl)
        self.activity(activity_type="field_event", activity_name_snapshot="Conference")
        self.client.force_login(self.pl)
        response = self.client.get(
            "/work-plan", {"fy": "2026", "view": "fy", "scope": "team"}
        )
        self.assertContains(response, 'href="#work-plan-detail"')
        self.assertContains(response, 'id="work-plan-detail"')
        self.assertContains(response, "See Details")
        self.assertContains(response, "School Activities")
        self.assertContains(response, "Non-School Activities")
        self.assertContains(response, 'name="scope" value="team"')
        for tab in response.context["view_tabs"]:
            self.assertIn("scope=team", tab["url"])

    def test_export_uses_the_same_scope_totals_and_detail_groups(self):
        self.activity(owner=self.pl, amount=42)
        self.activity(amount=75)
        self.client.force_login(self.pl)
        for scope, total, count in (("own", 42, 1), ("team", 117, 2)):
            response = self.client.get(
                "/work-plan/export.xlsx", {"fy": "2026", "view": "fy", "scope": scope}
            )
            self.assertEqual(response.status_code, 200)
            workbook = load_workbook(BytesIO(response.content), read_only=True)
            summary = list(workbook["Plan Summary"].values)
            self.assertEqual(
                next(r for r in summary if r[0] == "Grand total"),
                ("Grand total", count, None, total),
            )
            self.assertEqual(workbook["Work Plan"].max_row, count + 1)
            self.assertEqual(workbook["Work Plan"]["A2"].value, "School Activities")
